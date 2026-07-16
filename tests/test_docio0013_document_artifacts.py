from __future__ import annotations

import asyncio
import io
import json
import multiprocessing
import sys
import types
import uuid
from pathlib import Path

import pytest
from fastapi import FastAPI, Header, HTTPException
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool


ROOT = Path(__file__).resolve().parents[1]
app_alias = sys.modules.get("app")
if app_alias is None:
    app_alias = types.ModuleType("app")
    app_alias.__path__ = [str(ROOT / "app"), str(ROOT)]  # type: ignore[attr-defined]
    sys.modules["app"] = app_alias

from app.db import Base, get_db
from app.models import AuditLog, File, FileChunk, FileText, Message, Thread, ThreadMember
from app.routes import document_artifacts as route_module
from app.routes.document_artifacts import (
    DocumentArtifactRouterDeps,
    _read_limited_json_body,
    build_document_artifacts_router,
)
from app.services.document_artifact_service import (
    DocumentArtifactConcurrencyError,
    DocumentArtifactLimitError,
    DocumentArtifactLimits,
    DocumentArtifactTimeoutError,
    _active_generation_slots,
    _release_generation_slot,
    _try_acquire_generation_slot,
    generate_document_artifact,
    generate_document_artifact_isolated,
)
from app.services.document_governance_service import (
    DocioCoordinationUnavailable,
    DocioQuotaExceeded,
    DocioQuotaPolicy,
    distributed_lock,
    enforce_prospective_quota,
)


@pytest.fixture()
def database():
    engine = create_engine(
        "sqlite+pysqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    try:
        yield factory
    finally:
        Base.metadata.drop_all(engine)
        engine.dispose()


@pytest.fixture()
def client(database, monkeypatch):
    def override_db():
        db = database()
        try:
            yield db
        finally:
            db.close()

    def current_user(
        x_test_user: str = Header(default="user-a"),
        x_test_role: str = Header(default="user"),
        x_test_org: str = Header(default="org-a"),
    ):
        return {
            "sub": x_test_user,
            "role": x_test_role,
            "org_slug": x_test_org,
            "name": x_test_user,
            "email": f"{x_test_user}@example.test",
        }

    def get_request_org(user, x_org_slug):
        return str(x_org_slug or user["org_slug"])

    def check_member(db: Session, org: str, thread_id: str, user_id: str):
        return db.execute(
            select(ThreadMember).where(
                ThreadMember.org_slug == org,
                ThreadMember.thread_id == thread_id,
                ThreadMember.user_id == user_id,
            )
        ).scalar_one_or_none()

    def require_member(db: Session, org: str, thread_id: str, user_id: str):
        row = check_member(db, org, thread_id, user_id)
        if row is None:
            raise HTTPException(status_code=403, detail="thread_access_denied")
        return row

    app = FastAPI()
    app.include_router(
        build_document_artifacts_router(
            DocumentArtifactRouterDeps(
                get_current_user=current_user,
                get_request_org=get_request_org,
                require_thread_member=require_member,
                check_thread_member=check_member,
                new_id=lambda: uuid.uuid4().hex,
                now_ts=lambda: 1_720_000_000,
                logger=route_module.logging.getLogger("docio-test"),
                session_factory=database,
            )
        )
    )
    app.dependency_overrides[get_db] = override_db

    # Route transaction tests do not need to exercise process spawning repeatedly.
    monkeypatch.setattr(
        route_module,
        "generate_document_artifact_isolated",
        lambda payload, limits=None: generate_document_artifact(
            payload,
            limits=limits,
        ),
    )

    with TestClient(app) as test_client:
        yield test_client


def _seed(database):
    with database() as db:
        db.add_all(
            [
                Thread(
                    id="thread-a",
                    org_slug="org-a",
                    title="A",
                    created_at=1_719_999_000,
                ),
                Thread(
                    id="thread-b",
                    org_slug="org-a",
                    title="B",
                    created_at=1_719_999_000,
                ),
                ThreadMember(
                    id="member-a",
                    org_slug="org-a",
                    thread_id="thread-a",
                    user_id="user-a",
                    role="member",
                    created_at=1_719_999_000,
                ),
                File(
                    id="file-a",
                    org_slug="org-a",
                    thread_id="thread-a",
                    uploader_id="uploader-a",
                    uploader_name="Uploader",
                    uploader_email="uploader@example.test",
                    filename='relatório"\r\nmalicioso.txt',
                    original_filename="relatorio.txt",
                    origin="upload",
                    scope_thread_id="thread-a",
                    mime_type="text/plain",
                    size_bytes=5,
                    content=b"hello",
                    extraction_failed=False,
                    is_institutional=False,
                    created_at=1_719_999_000,
                ),
                File(
                    id="file-b",
                    org_slug="org-b",
                    thread_id="thread-z",
                    uploader_id="user-b",
                    uploader_name="B",
                    uploader_email="b@example.test",
                    filename="secret.txt",
                    original_filename="secret.txt",
                    origin="upload",
                    scope_thread_id="thread-z",
                    mime_type="text/plain",
                    size_bytes=6,
                    content=b"secret",
                    extraction_failed=False,
                    is_institutional=False,
                    created_at=1_719_999_000,
                ),
                File(
                    id="file-c",
                    org_slug="org-a",
                    thread_id="thread-b",
                    uploader_id="user-b",
                    uploader_name="B",
                    uploader_email="b@example.test",
                    filename="other-thread.txt",
                    original_filename="other-thread.txt",
                    origin="upload",
                    scope_thread_id="thread-b",
                    mime_type="text/plain",
                    size_bytes=12,
                    content=b"other-thread",
                    extraction_failed=False,
                    is_institutional=False,
                    created_at=1_719_999_000,
                ),
            ]
        )
        db.commit()


def test_formula_injection_and_zero_false_are_preserved():
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    artifact = generate_document_artifact(
        {
            "format": "xlsx",
            "title": "=1+1",
            "content": "+cmd",
            "rows": [[0, False, "@x", "-y"]],
        }
    )
    workbook = load_workbook(io.BytesIO(artifact.content), data_only=False)
    sheet = workbook.active

    assert sheet["A1"].value == "'=1+1"
    assert sheet["A2"].value == "'+cmd"
    assert sheet["A4"].value == "0"
    assert sheet["B4"].value == "False"
    assert sheet["C4"].value == "'@x"
    assert sheet["D4"].value == "'-y"
    assert all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)


def test_trusted_limits_reject_oversized_rows():
    limits = DocumentArtifactLimits(
        max_request_bytes=1024,
        max_rows=1,
        max_columns=2,
        max_cell_chars=10,
        max_total_cells=2,
        max_artifact_bytes=1024 * 1024,
        generation_timeout_seconds=2.0,
    )
    with pytest.raises(DocumentArtifactLimitError, match="max_rows_exceeded"):
        generate_document_artifact(
            {"format": "md", "title": "x", "rows": [["a"], ["b"]]},
            limits=limits,
        )


@pytest.mark.filterwarnings(
    "ignore:This process .* is multi-threaded, use of fork.*:DeprecationWarning"
)
def test_isolated_timeout_leaves_no_live_docio_worker(monkeypatch):
    # The hosted test interpreter injects a startup hook incompatible with
    # multiprocessing spawn. Production keeps the safer spawn default.
    monkeypatch.setenv("ORKIO_DOCIO_PROCESS_START_METHOD", "fork")
    rows = [["x" * 100, str(index)] for index in range(2_000)]
    limits = DocumentArtifactLimits(
        max_request_bytes=8 * 1024 * 1024,
        max_rows=2_000,
        max_columns=80,
        max_cell_chars=4_000,
        max_total_cells=50_000,
        max_artifact_bytes=10 * 1024 * 1024,
        generation_timeout_seconds=0.05,
    )
    with pytest.raises(DocumentArtifactTimeoutError):
        generate_document_artifact_isolated(
            {
                "format": "pdf",
                "title": "slow",
                "content": "body " * 20_000,
                "rows": rows,
            },
            limits=limits,
            timeout_seconds=0.001,
        )
    assert not [
        child
        for child in multiprocessing.active_children()
        if child.name == "orkio-docio-generator" and child.is_alive()
    ]
    assert _active_generation_slots() == 0


def test_generation_concurrency_slot_fails_closed_and_releases(monkeypatch):
    monkeypatch.setenv("ORKIO_DOCIO_MAX_CONCURRENT_GENERATIONS", "1")
    assert _try_acquire_generation_slot() is True
    assert _active_generation_slots() == 1
    try:
        with pytest.raises(
            DocumentArtifactConcurrencyError,
            match="concurrency_limit_reached",
        ):
            generate_document_artifact_isolated(
                {"format": "md", "title": "blocked"}
            )
        assert _active_generation_slots() == 1
    finally:
        _release_generation_slot()
    assert _active_generation_slots() == 0


def test_request_body_limit_is_enforced_before_schema(client, monkeypatch, database):
    monkeypatch.setenv("ORKIO_DOCIO_MAX_REQUEST_BYTES", "16384")
    payload = {
        "format": "md",
        "title": "x",
        "content": "a" * 20_000,
    }
    response = client.post(
        "/api/document-artifacts/generate",
        json=payload,
    )
    assert response.status_code == 413
    with database() as db:
        assert db.execute(select(File)).scalars().all() == []
        audit = db.execute(select(AuditLog)).scalars().one()
        assert audit.status_code == 413



def test_streaming_body_cap_works_without_or_with_misleading_content_length():
    def make_request(headers):
        chunks = iter(
            [
                {"type": "http.request", "body": b'{"content":"', "more_body": True},
                {"type": "http.request", "body": b"x" * 64, "more_body": True},
                {"type": "http.request", "body": b'"}', "more_body": False},
            ]
        )

        async def receive():
            try:
                return next(chunks)
            except StopIteration:
                return {"type": "http.request", "body": b"", "more_body": False}

        scope = {
            "type": "http",
            "asgi": {"version": "3.0"},
            "http_version": "1.1",
            "method": "POST",
            "scheme": "http",
            "path": "/api/document-artifacts/generate",
            "raw_path": b"/api/document-artifacts/generate",
            "query_string": b"",
            "headers": headers,
            "client": ("test", 123),
            "server": ("test", 80),
        }
        return route_module.Request(scope, receive)

    for headers in (
        [],
        [(b"content-length", b"1")],
    ):
        with pytest.raises(HTTPException) as exc_info:
            asyncio.run(
                _read_limited_json_body(
                    make_request(headers),
                    max_bytes=32,
                )
            )
        assert exc_info.value.status_code == 413
        assert exc_info.value.detail == "request_body_too_large"


def test_public_schema_rejects_runtime_identity_fields(client, database):
    response = client.post(
        "/api/document-artifacts/generate",
        json={
            "format": "md",
            "title": "x",
            "agent_name": "Orion",
            "execution_id": "spoof",
            "resolved_agent": "Orion",
            "limits": {"max_rows": 999999},
        },
    )
    assert response.status_code == 422
    with database() as db:
        assert db.execute(select(File)).scalars().all() == []


def test_generation_commits_file_index_event_and_audit_atomically(
    client,
    database,
):
    _seed(database)
    response = client.post(
        "/api/document-artifacts/generate",
        headers={"x-test-user": "user-a", "x-test-org": "org-a"},
        json={
            "format": "md",
            "title": "Plano",
            "content": "Conteúdo",
            "thread_id": "thread-a",
            "requested_agent_hint": "Orion",
        },
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["audit_persisted"] is True
    assert body["thread_event_persisted"] is True
    assert body["resolved_agent"] is None
    assert body["authorship_claim"] == "user_or_system"

    with database() as db:
        file_row = db.get(File, body["file_id"])
        assert file_row is not None
        assert file_row.org_slug == "org-a"
        assert db.execute(
            select(FileText).where(FileText.file_id == body["file_id"])
        ).scalar_one()
        assert db.execute(
            select(FileChunk).where(FileChunk.file_id == body["file_id"])
        ).scalars().all()
        event = db.execute(
            select(Message).where(Message.thread_id == "thread-a")
        ).scalar_one()
        assert event.agent_name is None
        assert '"authorship_claim": "user_or_system"' in event.content
        audit = db.execute(
            select(AuditLog).where(
                AuditLog.action == "document_artifact.generated"
            )
        ).scalar_one()
        assert audit.status_code == 200


def test_generation_rolls_back_when_success_audit_cannot_be_staged(
    client,
    database,
    monkeypatch,
):
    _seed(database)

    def fail_success_audit(*args, **kwargs):
        raise RuntimeError("audit unavailable")

    monkeypatch.setattr(route_module, "stage_audit", fail_success_audit)
    response = client.post(
        "/api/document-artifacts/generate",
        headers={"x-test-user": "user-a", "x-test-org": "org-a"},
        json={
            "format": "md",
            "title": "Plano",
            "thread_id": "thread-a",
        },
    )
    assert response.status_code == 503
    with database() as db:
        assert db.execute(
            select(File).where(File.origin == "generated")
        ).scalars().all() == []
        assert db.execute(select(FileText)).scalars().all() == []
        assert db.execute(select(FileChunk)).scalars().all() == []
        # The separate failure audit remains observable.
        failures = db.execute(
            select(AuditLog).where(
                AuditLog.action == "document_artifact.generate.failed"
            )
        ).scalars().all()
        assert len(failures) == 1


def test_download_acl_and_hardened_headers(client, database):
    _seed(database)

    member = client.get(
        "/api/files/file-a/download",
        headers={"x-test-user": "user-a", "x-test-org": "org-a"},
    )
    assert member.status_code == 200
    assert member.content == b"hello"
    disposition = member.headers["content-disposition"]
    assert "\r" not in disposition and "\n" not in disposition
    assert "filename*=UTF-8''" in disposition
    assert member.headers["cache-control"] == "private, no-store"
    assert member.headers["x-content-type-options"] == "nosniff"

    uploader = client.get(
        "/api/files/file-a/download",
        headers={"x-test-user": "uploader-a", "x-test-org": "org-a"},
    )
    assert uploader.status_code == 200

    admin = client.get(
        "/api/files/file-a/download",
        headers={
            "x-test-user": "admin-a",
            "x-test-role": "admin",
            "x-test-org": "org-a",
        },
    )
    assert admin.status_code == 200

    unauthorized = client.get(
        "/api/files/file-a/download",
        headers={"x-test-user": "user-x", "x-test-org": "org-a"},
    )
    assert unauthorized.status_code == 403

    cross_thread = client.get(
        "/api/files/file-c/download",
        headers={"x-test-user": "user-a", "x-test-org": "org-a"},
    )
    assert cross_thread.status_code == 403

    cross_tenant = client.get(
        "/api/files/file-b/download",
        headers={
            "x-test-user": "admin-a",
            "x-test-role": "admin",
            "x-test-org": "org-a",
        },
    )
    assert cross_tenant.status_code == 404



def test_generation_blocks_cross_thread_membership(client, database):
    _seed(database)
    response = client.post(
        "/api/document-artifacts/generate",
        headers={"x-test-user": "user-a", "x-test-org": "org-a"},
        json={
            "format": "md",
            "title": "Blocked",
            "thread_id": "thread-b",
        },
    )
    assert response.status_code == 403
    with database() as db:
        assert db.execute(
            select(File).where(File.origin == "generated")
        ).scalars().all() == []
        rejected = db.execute(
            select(AuditLog).where(
                AuditLog.action == "document_artifact.generate.rejected"
            )
        ).scalars().one()
        assert rejected.status_code == 403


def test_prospective_quota_counts_new_bytes(database):
    with database() as db:
        db.add(
            File(
                id="existing",
                org_slug="org-a",
                thread_id=None,
                uploader_id="user-a",
                uploader_name="A",
                uploader_email="a@example.test",
                filename="existing.md",
                original_filename="existing.md",
                origin="generated",
                scope_thread_id=None,
                mime_type="text/markdown",
                size_bytes=90,
                content=b"x" * 90,
                extraction_failed=False,
                is_institutional=False,
                created_at=1_720_000_000,
            )
        )
        db.commit()

        policy = DocioQuotaPolicy(
            user_count_per_hour=10,
            tenant_count_per_hour=10,
            user_bytes_per_hour=100,
            tenant_bytes_per_hour=100,
            tenant_bytes_per_month=100,
        )
        with pytest.raises(DocioQuotaExceeded, match="user_bytes_hour"):
            enforce_prospective_quota(
                db,
                org="org-a",
                user_id="user-a",
                prospective_bytes=11,
                now_ts=1_720_000_001,
                policy=policy,
            )


def test_reindex_is_admin_only_and_rate_limited(client, database, monkeypatch):
    _seed(database)
    denied = client.post(
        "/api/files/file-a/reindex",
        headers={"x-test-user": "user-a", "x-test-org": "org-a"},
    )
    assert denied.status_code == 403

    monkeypatch.setenv("ORKIO_DOCIO_REINDEX_MIN_INTERVAL_SECONDS", "300")
    first = client.post(
        "/api/files/file-a/reindex",
        headers={
            "x-test-user": "admin-a",
            "x-test-role": "admin",
            "x-test-org": "org-a",
        },
    )
    assert first.status_code == 200, first.text

    second = client.post(
        "/api/files/file-a/reindex",
        headers={
            "x-test-user": "admin-a",
            "x-test-role": "admin",
            "x-test-org": "org-a",
        },
    )
    assert second.status_code == 429
    with database() as db:
        actions = [
            row.action for row in db.execute(select(AuditLog)).scalars().all()
        ]
        assert "document_artifact.reindex.completed" in actions
        assert "document_artifact.reindex.rate_limited" in actions


def test_readonly_context_has_no_lazy_reindex_static_contract():
    source = (ROOT / "services" / "document_context_service.py").read_text(
        encoding="utf-8"
    )
    assert "_maybe_reindex_file_from_stored_content" not in source
    assert "RTB09_LAZY_FILE_REINDEX" not in source


def test_main_registers_router_without_inline_document_endpoints():
    source = (ROOT / "main.py").read_text(encoding="utf-8")
    assert "build_document_artifacts_router" in source
    assert "def generate_document_artifact_endpoint" not in source
    assert '@app.get("/api/documents/thread-context")' not in source
    assert '@app.get("/api/files/{file_id}/download")' not in source

def test_postgresql_advisory_lock_uses_dedicated_connection():
    calls = []

    class FakeResult:
        def scalar(self):
            return True

    class FakeConnection:
        def execute(self, statement, params):
            calls.append((str(statement), dict(params)))
            return FakeResult()

        def rollback(self):
            calls.append(("rollback", {}))

        def close(self):
            calls.append(("close", {}))

    class FakeEngine:
        dialect = types.SimpleNamespace(name="postgresql")

        def connect(self):
            calls.append(("connect", {}))
            return FakeConnection()

    class FakeSession:
        def get_bind(self):
            return FakeEngine()

    with distributed_lock(
        FakeSession(),  # type: ignore[arg-type]
        namespace="docio_test",
        value="org-a",
    ):
        calls.append(("inside", {}))

    statements = [item[0] for item in calls]
    assert any("pg_try_advisory_lock" in item for item in statements)
    assert "inside" in statements
    assert any("pg_advisory_unlock" in item for item in statements)
    assert statements[-1] == "close"


def test_distributed_lock_fails_closed_on_unsupported_dialect():
    class UnsupportedBind:
        dialect = types.SimpleNamespace(name="mysql")

    class FakeSession:
        def get_bind(self):
            return UnsupportedBind()

    with pytest.raises(
        DocioCoordinationUnavailable,
        match="distributed_coordination_requires_postgresql:mysql",
    ):
        with distributed_lock(
            FakeSession(),  # type: ignore[arg-type]
            namespace="docio_test",
            value="org-a",
        ):
            raise AssertionError("unreachable")
